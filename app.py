"""
app.py — DBS Check Service (NextStep SaaS)
Full integration version: auth/session acceptance, embedded child-task dispatcher,
identity chain, isolated storage, ownership checks, token tracking.
"""
import sys, asyncio, anyio, logging, json, os, io, csv, re, secrets
import datetime, shutil, uuid, zipfile, time
from pathlib import Path
from typing import Dict, Any, Optional, Tuple, List
from zoneinfo import ZoneInfo

if sys.platform.startswith("win"):
    asyncio.set_event_loop_policy(asyncio.WindowsProactorEventLoopPolicy())

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(name)s — %(message)s")
log = logging.getLogger(__name__)

from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Request
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse, FileResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.background import BackgroundTask
from starlette.middleware.base import BaseHTTPMiddleware

import fitz
import pdfplumber

try:
    from google import genai
    from google.genai import types
except Exception:
    genai = None; types = None

APP_DIR      = os.path.dirname(os.path.abspath(__file__))
REDIS_URL    = os.environ.get("REDIS_URL", "redis://localhost:6379")
STORAGE_ROOT = Path("/tmp/nextstep")
STORAGE_ROOT.mkdir(parents=True, exist_ok=True)
BACKEND_VALIDATE_URL = os.environ.get("BACKEND_VALIDATE_URL", "https://nextstep-backend-e75l.onrender.com/api/validate-session")
APP_DASHBOARD_URL = os.environ.get("APP_DASHBOARD_URL", "https://nextstep-backend-e75l.onrender.com/dashboard")
APP_LOGIN_URL = os.environ.get("APP_LOGIN_URL", "https://nextstep-backend-e75l.onrender.com/login")
MAX_CONCURRENT_TASKS = max(1, int(os.environ.get("MAX_CONCURRENT_TASKS", "1")))
WORKER_POLL_INTERVAL = max(1, int(os.environ.get("WORKER_POLL_INTERVAL", "2")))

JOB_PREFIX = "nextstep:dbs:job:"
OWNER_PREFIX = "nextstep:dbs:owner:"
CHILD_PREFIX = "nextstep:dbs:child:"
ACTIVE_CHILDREN_KEY = "nextstep:dbs:children:active"

_redis = None
_dispatcher_started = False
_dispatcher_lock = __import__('threading').Lock()
_local_active = set()
_local_active_lock = __import__('threading').Lock()
_last_parent_pointer = 0

class PersistNextStepTokenMiddleware(BaseHTTPMiddleware):
    async def dispatch(self, request, call_next):
        response = await call_next(request)
        token = request.query_params.get("ns_token")
        if token:
            response.set_cookie(key="ns_token", value=token, httponly=True, samesite="lax", secure=True, max_age=60 * 60 * 8)
        return response

app = FastAPI(title="DBS Check — NextStep")
app.add_middleware(PersistNextStepTokenMiddleware)
app.mount("/static", StaticFiles(directory=os.path.join(APP_DIR, "static")), name="static")
templates = Jinja2Templates(directory=os.path.join(APP_DIR, "templates"))

# ── Gemini ────────────────────────────────────────────────────────────────────
def _env(name, default=None):
    v = os.getenv(name)
    return default if v is None or str(v).strip() == "" else v

GEMINI_API_KEY    = _env("GEMINI_API_KEY")
GEMINI_MODEL_FAST = _env("GEMINI_MODEL_FAST", "gemini-2.0-flash-001")
GEMINI_MODEL_STRONG = _env("GEMINI_MODEL_STRONG", "gemini-2.5-pro")

def get_gemini_client():
    if not GEMINI_API_KEY or genai is None: return None
    try: return genai.Client(api_key=GEMINI_API_KEY)
    except Exception: return None

GEMINI_CLIENT = get_gemini_client()

# ── Redis ─────────────────────────────────────────────────────────────────────
def get_redis():
    global _redis
    if _redis is None:
        try:
            import redis
            _redis = redis.Redis.from_url(REDIS_URL, decode_responses=False)
            _redis.ping()
        except Exception as e:
            log.error("[Redis] %s", e); _redis = None
    return _redis


def _job_key(job_id: str) -> str:
    return f"{JOB_PREFIX}{job_id}"


def _owner_key(job_id: str) -> str:
    return f"{OWNER_PREFIX}{job_id}"


def _child_key(child_id: str) -> str:
    return f"{CHILD_PREFIX}{child_id}"


def _parent_queue_key(job_id: str) -> str:
    return f"nextstep:dbs:parent:{job_id}:queue"


def _job_to_dict(raw):
    if not raw:
        return None
    try:
        if isinstance(raw, bytes):
            raw = raw.decode("utf-8")
        return json.loads(raw)
    except Exception:
        return None


def _jget(job_id):
    r = get_redis()
    if not r: return None
    return _job_to_dict(r.get(_job_key(job_id)))


def _jset(job_id, state):
    r = get_redis()
    if not r: return
    try: r.setex(_job_key(job_id), 60 * 60 * 8, json.dumps(state))
    except: pass


def _owner_set(job_id, tenant_id):
    r = get_redis()
    if not r: return
    try: r.setex(_owner_key(job_id), 60 * 60 * 8, str(tenant_id))
    except: pass


def _owner_get(job_id):
    r = get_redis()
    if not r: return None
    try:
        v = r.get(_owner_key(job_id))
        return int(v) if v else None
    except: return None




def _row_is_billable(row: Dict[str, Any]) -> bool:
    if not isinstance(row, dict):
        return False
    status = str(row.get("status") or "").strip()
    has_pdf = bool((row.get("pdf_filename") or "").strip() or (row.get("pdf_url") or "").strip())
    return status in ("clear", "needs_review") and has_pdf and not bool(row.get("billing_waived"))

def _child_get(child_id: str):
    r = get_redis()
    if not r:
        return None
    return _job_to_dict(r.get(_child_key(child_id)))


def _child_set(child_id: str, payload: dict):
    r = get_redis()
    if not r:
        return
    r.setex(_child_key(child_id), 60 * 60 * 8, json.dumps(payload))

# ── Auth ──────────────────────────────────────────────────────────────────────
def _validate_via_backend(token: str):
    if not token:
        return None
    try:
        import requests
        resp = requests.get(BACKEND_VALIDATE_URL, params={"token": token}, timeout=8)
        if resp.status_code != 200:
            return None
        data = resp.json()
        if not data.get("valid"):
            return None
        user = data.get("user") or {}
        tenant = data.get("tenant") or {}
        return {
            "user_id": user.get("id"),
            "tenant_id": tenant.get("id"),
            "role": user.get("role", "admin"),
            "email": user.get("email"),
            "name": user.get("name"),
        }
    except Exception as e:
        log.warning("[Auth backend] %s", e)
        return None


def _get_ctx(request: Request):
    token = (request.headers.get("X-NextStep-Token")
             or request.cookies.get("ns_token")
             or request.query_params.get("ns_token") or "")
    if not token: return None
    ctx = _validate_via_backend(token)
    if ctx:
        return ctx
    try:
        import db; return db.validate_user_token(token)
    except Exception as e:
        log.warning("[Auth db] %s", e); return None


def _auth(request: Request):
    ctx = _get_ctx(request)
    if not ctx: raise HTTPException(401, f"Not authenticated. Please log in at {APP_LOGIN_URL}")
    return ctx

# ── Storage ───────────────────────────────────────────────────────────────────
def _storage(tenant_id, user_id, job_id):
    p = STORAGE_ROOT / str(tenant_id) / str(user_id) / job_id
    p.mkdir(parents=True, exist_ok=True)
    return p

# ── Helpers (kept from original) ─────────────────────────────────────────────
def normalize_ws(s):
    return re.sub(r"\s+", " ", (s or "").strip())

MONTHS = {
    "JANUARY":"01","FEBRUARY":"02","MARCH":"03","APRIL":"04","MAY":"05","JUNE":"06",
    "JULY":"07","AUGUST":"08","SEPTEMBER":"09","OCTOBER":"10","NOVEMBER":"11","DECEMBER":"12",
    "JAN":"01","FEB":"02","MAR":"03","APR":"04","JUN":"06","JUL":"07","AUG":"08",
    "SEP":"09","SEPT":"09","OCT":"10","NOV":"11","DEC":"12",
}
MONTH_ALIASES = {
    "JANUARI":"JANUARY","FEBRUARI":"FEBRUARY","MARH":"MARCH","APRL":"APRIL",
    "JUNE":"JUNE","JULY":"JULY","AUGI":"AUG","AUGL":"AUG","AUC":"AUG",
    "SEPTEM8ER":"SEPTEMBER","SEPTEMPER":"SEPTEMBER","0CT":"OCT","NOVEM8ER":"NOVEMBER",
    "DECEM8ER":"DECEMBER",
}


def normalize_month_value(value):
    s = str(value or "").strip()
    if not s:
        return ""
    if s.isdigit():
        n = int(s)
        return f"{n:02d}" if 1 <= n <= 12 else ""
    token = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    if not token:
        return ""
    token = token.replace("0", "O").replace("1", "I").replace("5", "S").replace("8", "B")
    token = MONTH_ALIASES.get(token, token)
    if token in MONTHS:
        return MONTHS[token]
    if len(token) >= 3:
        short = token[:3]
        if short in MONTHS:
            return MONTHS[short]
        import difflib
        match = difflib.get_close_matches(short, [k for k in MONTHS if len(k) == 3], n=1, cutoff=0.66)
        if match:
            return MONTHS[match[0]]
    match = None
    try:
        import difflib
        match = difflib.get_close_matches(token, [k for k in MONTHS if len(k) > 3], n=1, cutoff=0.72)
    except Exception:
        match = None
    return MONTHS.get(match[0], "") if match else ""


def normalize_dob_parts(dd, mm, yy):
    dd = str(dd or "").strip()
    mm = normalize_month_value(mm)
    yy = str(yy or "").strip()
    if dd.isdigit():
        dd = dd.zfill(2)
    if len(yy) == 2 and yy.isdigit():
        yy = ("19" if int(yy) > 30 else "20") + yy
    return dd, mm, yy


def parse_uk_date_words(date_str):
    if not date_str: return None
    s = normalize_ws(date_str).upper()
    m = re.search(r"\b(\d{1,2})\s+([A-Z]{3,9})\s+(\d{4})\b", s)
    if not m: return None
    dd = m.group(1).zfill(2); mon = normalize_month_value(m.group(2)); yyyy = m.group(3)
    return (dd, mon, yyyy) if mon else None

def parse_ddmmyyyy(date_str):
    if not date_str: return None
    m = re.search(r"\b(\d{1,2})[\/\.-](\d{1,2})[\/\.-](\d{2,4})\b", normalize_ws(date_str))
    if not m: return None
    dd = m.group(1).zfill(2); mm = normalize_month_value(m.group(2)); yy = m.group(3)
    if len(yy) == 2: yy = ("19" if int(yy) > 30 else "20") + yy
    return (dd, mm, yy) if mm else None

def validate_cert_number(s):
    if not s: return None
    digits = re.sub(r"\D", "", s)
    return digits if 10 <= len(digits) <= 14 else None

def score_cert_number(val):
    digits = re.sub(r"\D", "", val or "")
    if not digits: return 0
    return min(100, 80 + (20 if validate_cert_number(digits) else 0)) if 10 <= len(digits) <= 14 else 30

def score_surname(val, *, source=""):
    s = (val or "").strip()
    if not s: return 0
    cleaned = re.sub(r"[^A-Za-z\-\s]", "", s).strip()
    if not cleaned: return 10
    length = len(cleaned.replace(" ", ""))
    score = 35 if length < 2 else (70 if length < 4 else 85)
    if cleaned != s: score -= 10
    if (source or "").lower().startswith("pdf"): score = min(100, score + 5)
    return max(0, min(100, score))

def score_dob(dd, mm, yyyy, *, source=""):
    dd, mm, yyyy = str(dd or "").strip(), str(mm or "").strip(), str(yyyy or "").strip()
    if not (dd or mm or yyyy): return 0
    if not (dd and mm and yyyy): return 45
    try:
        datetime.date(int(yyyy), int(mm), int(dd))
        score = 95
        if (source or "").lower().startswith("pdf"): score = min(100, score+5)
        return score
    except: return 20

def _safe_filename(name, default):
    name = normalize_ws(name).strip() or default
    name = re.sub(r'[\\/:"*?<>|]+', "-", name)
    return re.sub(r"\s+", " ", name).strip()

def _uk_checked_date():
    return datetime.datetime.now(tz=ZoneInfo("Europe/London")).strftime("%d.%m.%Y")

def _dmy(dd, mm, yy):
    dd, mm, yy = str(dd or "").strip(), str(mm or "").strip(), str(yy or "").strip()
    return f"{dd.zfill(2)}/{mm.zfill(2)}/{yy}" if (dd and mm and yy) else ""

# ── PDF extraction helpers (kept from original) ───────────────────────────────
def extract_text_from_pdf(pdf_bytes, max_pages=2):
    try:
        with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
            return "\n".join([p.extract_text() or "" for p in pdf.pages[:max_pages]]).strip()
    except: return ""

def pdf_to_images_bytes(pdf_bytes, max_pages=1, dpi=240):
    images = []
    if not pdf_bytes or len(pdf_bytes) < 100: raise ValueError("File too small.")
    if not pdf_bytes.lstrip().startswith(b"%PDF"): raise ValueError("Not a PDF.")
    try: doc = fitz.open(stream=pdf_bytes, filetype="pdf")
    except Exception as e: raise ValueError(f"Cannot open PDF: {e}")
    try:
        if getattr(doc, "is_encrypted", False):
            try: doc.authenticate("")
            except: pass
        if getattr(doc, "page_count", 0) <= 0: return images
        page = doc.load_page(0)
        zoom = dpi / 72.0; mat = fitz.Matrix(zoom, zoom); rect = page.rect
        for x0,y0,x1,y1 in [(0,0,1,.32),(0,.28,1,.72),(0,.68,1,1)]:
            clip = fitz.Rect(rect.x0+rect.width*x0, rect.y0+rect.height*y0,
                             rect.x0+rect.width*x1, rect.y0+rect.height*y1)
            images.append(page.get_pixmap(matrix=mat, clip=clip, alpha=False).tobytes("png"))
        images.append(page.get_pixmap(matrix=mat, alpha=False).tobytes("png"))
        return images
    finally:
        try: doc.close()
        except: pass

def extract_fields_from_text(text):
    t = normalize_ws(text)
    out = {"certificate_number": None, "surname": None, "dob": None}
    m = re.search(r"Certificate\s*Number[:\s]*([0-9\s]{8,20})", t, re.IGNORECASE)
    if m: out["certificate_number"] = validate_cert_number(m.group(1))
    m = re.search(r"Surname[:\s]*([A-Z'\-\s]{2,40})", t, re.IGNORECASE)
    if m: out["surname"] = normalize_ws(m.group(1)).upper()
    m = re.search(r"Date\s*of\s*Birth[:\s]*([0-9]{1,2}\s+[A-Za-z]{3,9}\s+[0-9]{4}|\d{1,2}[\/\.-]\d{1,2}[\/\.-]\d{2,4})", t, re.IGNORECASE)
    if m:
        parts = parse_uk_date_words(m.group(1)) or parse_ddmmyyyy(m.group(1))
        if parts: out["dob"] = {"dd": parts[0], "mm": parts[1], "yyyy": parts[2]}
    return out

VISION_PROMPT = """Extract from UK DBS Certificate. Return STRICT JSON:
{"certificate_number":{"value":"","confidence":0.0},"surname":{"value":"","confidence":0.0},
"forename":{"value":"","confidence":0.0},"dob":{"day":"","month":"","year":"","confidence":0.0},
"issue_date":{"day":"","month":"","year":"","confidence":0.0}}
Certificate number = digits only. Do NOT use Print Date for issue_date."""

def _parse_json_response(text):
    if not text: return {}
    text = re.sub(r"^```(json)?", "", text.strip(), flags=re.IGNORECASE).strip()
    text = re.sub(r"```$", "", text).strip()
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m: return {}
    try: return json.loads(m.group(0))
    except: return {}

def gemini_vision_extract_images(images):
    if GEMINI_CLIENT is None or not images: return {}
    def _call(model):
        parts = [types.Part.from_text(text=VISION_PROMPT)]
        for b, mime in images: parts.append(types.Part.from_bytes(data=b, mime_type=mime))
        resp = GEMINI_CLIENT.models.generate_content(model=model, contents=[types.Content(role="user", parts=parts)])
        data = _parse_json_response(getattr(resp, "text", None) or "")
        if isinstance(data, dict): data["_model"] = model
        return data if isinstance(data, dict) else {}

    def _missing(d):
        if not d: return True
        def _v(x): return str(x.get("value","") if isinstance(x, dict) else x or "")
        cert_ok = bool(validate_cert_number(_v(d.get("certificate_number",""))))
        sn_ok = bool(normalize_ws(_v(d.get("surname",""))).strip())
        dob = d.get("dob") or {}
        dob_ok = bool(str(dob.get("day","")).strip() and str(dob.get("month","")).strip() and str(dob.get("year","")).strip()) if isinstance(dob, dict) else False
        return not (cert_ok and sn_ok and dob_ok)

    data = {}
    try: data = _call(GEMINI_MODEL_FAST)
    except: data = {}
    if _missing(data):
        try: data = _call(GEMINI_MODEL_STRONG)
        except: pass

    def _gv(k): v = data.get(k); return str(v.get("value","") if isinstance(v, dict) else v or "")
    def _gc(k):
        v = data.get(k)
        try: return float(v.get("confidence", 0) if isinstance(v, dict) else 0)
        except: return 0.0
    def _pct(f):
        try: return int(round(max(0.0,min(1.0,float(f or 0)))*100))
        except: return 0

    out = {
        "certificate_number": validate_cert_number(_gv("certificate_number")),
        "surname": normalize_ws(_gv("surname")).upper() or None,
        "forename": normalize_ws(_gv("forename")).upper() or None,
        "dob": None, "issue_date": None,
        "confidence": {"certificate_number": _pct(_gc("certificate_number")),
                       "surname": _pct(_gc("surname")), "dob": _pct(_gc("dob")),
                       "issue_date": _pct(_gc("issue_date"))},
        "_model": str(data.get("_model",""))
    }
    def _parse_date_obj(key):
        obj = data.get(key)
        if isinstance(obj, dict):
            dd = str(obj.get("day","")).zfill(2) if str(obj.get("day","")).strip() else ""
            raw_mm = str(obj.get("month","")).strip()
            mm = normalize_month_value(raw_mm) if raw_mm else ""
            yy = str(obj.get("year","")).strip()
            if dd and mm and yy: return {"dd": dd, "mm": mm, "yyyy": yy}
        parts = parse_uk_date_words(_gv(key)) or parse_ddmmyyyy(_gv(key))
        return {"dd": parts[0], "mm": parts[1], "yyyy": parts[2]} if parts else None
    out["dob"] = _parse_date_obj("dob")
    out["issue_date"] = _parse_date_obj("issue_date")
    return out

# ── Spreadsheet parsing (kept from original) ──────────────────────────────────
def _norm_col(name):
    return re.sub(r"[^a-z0-9]+", "", (name or "").strip().lower())

def _parse_dob_value(v):
    if v is None: return ("","","")
    if isinstance(v, (datetime.date, datetime.datetime)):
        d = v.date() if isinstance(v, datetime.datetime) else v
        return (str(d.day).zfill(2), str(d.month).zfill(2), str(d.year))
    s = normalize_ws(str(v).strip())
    if not s: return ("","","")
    m = re.match(r"^(\d{4})[-/](\d{1,2})[-/](\d{1,2})$", s)
    if m: return (m.group(3).zfill(2), normalize_month_value(m.group(2)), m.group(1))
    m = re.match(r"^(\d{1,2})[-/](\d{1,2})[-/](\d{4})$", s)
    if m: return (m.group(1).zfill(2), normalize_month_value(m.group(2)), m.group(3))
    parts = parse_uk_date_words(s) or parse_ddmmyyyy(s)
    return parts if parts else ("","","")

def parse_csv_rows(content):
    text = None
    for enc in ("utf-8-sig","utf-8","latin-1"):
        try: text = content.decode(enc); break
        except: continue
    if text is None: raise ValueError("Cannot decode CSV.")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames: raise ValueError("CSV has no header.")
    cols = {_norm_col(c): c for c in reader.fieldnames}
    return _rows_from_dict_iter(reader, cols)

def parse_xlsx_rows(content):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb.active; rows = list(ws.iter_rows(values_only=True))
    if not rows: raise ValueError("Excel is empty.")
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    cols = {_norm_col(c): c for c in headers}
    dict_rows = [{h: rows[i][j] if j < len(rows[i]) else None for j,h in enumerate(headers) if h} for i in range(1, len(rows))]
    return _rows_from_dict_iter(dict_rows, cols)

def _rows_from_dict_iter(iterable, cols):
    def fp(keys):
        for k in keys:
            if _norm_col(k) in cols: return cols[_norm_col(k)]
        return None
    cert_col = fp(["certificate_number","certno","certnumber","certificate","cert"])
    sn_col   = fp(["surname","lastname","last_name"])
    fn_col   = fp(["forename","firstname","first_name","givenname"])
    dob_col  = fp(["dob","dateofbirth","dateofbirthddmmyyyy"])
    dd_col   = fp(["dobday","day","dd"]); mm_col = fp(["dobmonth","month","mm"]); yy_col = fp(["dobyear","year","yyyy","yy"])
    issue_col = fp(["issuedate","dateofissue","issue_date"])
    if not cert_col or not sn_col or not (dob_col or (dd_col and mm_col and yy_col)):
        raise ValueError("Spreadsheet must have Certificate No, Surname, and DOB columns.")
    out = []
    for d in iterable:
        if not d: continue
        cert = str(d.get(cert_col) or "").strip()
        sn   = str(d.get(sn_col) or "").strip()
        fn   = str(d.get(fn_col) or "").strip() if fn_col else ""
        if dob_col: dd,mm,yy = _parse_dob_value(d.get(dob_col))
        else:
            dd = str(d.get(dd_col) or "").strip().zfill(2) if str(d.get(dd_col) or "").strip() else ""
            mm = str(d.get(mm_col) or "").strip().zfill(2) if str(d.get(mm_col) or "").strip() else ""
            yy = str(d.get(yy_col) or "").strip()
        if not (cert or sn or (dd and mm and yy)): continue
        issue_dd = issue_mm = issue_yy = ""
        if issue_col:
            ip = parse_uk_date_words(str(d.get(issue_col) or "")) or parse_ddmmyyyy(str(d.get(issue_col) or ""))
            if ip: issue_dd, issue_mm, issue_yy = ip
        out.append({"certificate_number":cert,"surname":sn,"forename":fn,
                    "dob_day":dd,"dob_month":mm,"dob_year":yy,
                    "issue_day":issue_dd,"issue_month":issue_mm,"issue_year":issue_yy})
    return out

# ── Export helpers ────────────────────────────────────────────────────────────
def _csv_bytes(rows, columns):
    buf = io.StringIO(); w = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
    w.writeheader(); [w.writerow({c: r.get(c,"") for c in columns}) for r in rows]
    return buf.getvalue().encode("utf-8-sig")

def _xlsx_bytes(rows, columns):
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active; ws.append(columns)
    [ws.append([r.get(c,"") for c in columns]) for r in rows]
    out = io.BytesIO(); wb.save(out); return out.getvalue()

def _export_rows_results(payload):
    checked_date = (payload.get("checked_date") or "").strip()
    return [{
        "Forename": (r.get("forename") or "").strip(),
        "Surname": (r.get("surname") or "").strip(),
        "Certificate Number": (r.get("certificate_number") or "").strip(),
        "DOB": _dmy(r.get("dob_day"), r.get("dob_month"), r.get("dob_year")),
        "Issue Date": _dmy(r.get("issue_day"), r.get("issue_month"), r.get("issue_year")),
        "Status": (r.get("status") or "").strip(),
        "Checked Date": checked_date,
        "PDF Filename": (r.get("pdf_filename") or "").strip(),
        "Notes": (r.get("error") or r.get("notes") or "").strip(),
    } for r in (payload.get("rows") or [])]

def _export_rows_extract(items):
    return [{
        "Forename": (it.get("forename") or "").strip(),
        "Surname": (it.get("surname") or "").strip(),
        "Certificate Number": (it.get("certificate_number") or "").strip(),
        "DOB": _dmy(it.get("dob_day"), it.get("dob_month"), it.get("dob_year")),
        "Issue Date": _dmy(it.get("issue_day"), it.get("issue_month"), it.get("issue_year")),
        "PDF Filename": (it.get("original_filename") or "").strip(),
        "Notes": "",
    } for it in (items or [])]

# ── Routes ────────────────────────────────────────────────────────────────────
@app.on_event("startup")
def startup_event():
    global _dispatcher_started
    with _dispatcher_lock:
        if _dispatcher_started:
            return
        _dispatcher_started = True
        import threading
        t = threading.Thread(target=_dispatcher_loop, daemon=True, name="dbs-child-dispatcher")
        t.start()
        log.info("[Dispatcher] started with MAX_CONCURRENT_TASKS=%d", MAX_CONCURRENT_TASKS)


def _recover_stale_children():
    r = get_redis()
    if not r:
        return
    try:
        for key in r.scan_iter(match=f"{CHILD_PREFIX}*"):
            child = _job_to_dict(r.get(key))
            if not child:
                continue
            if child.get("status") == "running":
                child["status"] = "queued"
                _child_set(child["child_id"], child)
        r.delete(ACTIVE_CHILDREN_KEY)
    except Exception as e:
        log.warning("[Dispatcher] recovery failed: %s", e)


def _list_parents_with_queue() -> List[str]:
    r = get_redis()
    if not r:
        return []
    job_ids = []
    for key in r.scan_iter(match="nextstep:dbs:parent:*:queue"):
        key_s = key.decode() if isinstance(key, bytes) else str(key)
        parts = key_s.split(":")
        if len(parts) >= 5:
            job_id = parts[3]
            try:
                if r.llen(key_s) > 0:
                    job_ids.append(job_id)
            except Exception:
                continue
    job_ids.sort()
    return job_ids


def _active_count() -> int:
    with _local_active_lock:
        return len(_local_active)


def _register_active(child_id: str):
    r = get_redis()
    with _local_active_lock:
        _local_active.add(child_id)
    if r:
        r.sadd(ACTIVE_CHILDREN_KEY, child_id)


def _unregister_active(child_id: str):
    r = get_redis()
    with _local_active_lock:
        _local_active.discard(child_id)
    if r:
        r.srem(ACTIVE_CHILDREN_KEY, child_id)


def _pick_next_child() -> Optional[str]:
    global _last_parent_pointer
    r = get_redis()
    if not r:
        return None
    parents = _list_parents_with_queue()
    if not parents:
        return None
    if _last_parent_pointer >= len(parents):
        _last_parent_pointer = 0
    ordered = parents[_last_parent_pointer:] + parents[:_last_parent_pointer]
    for job_id in ordered:
        qkey = _parent_queue_key(job_id)
        child_id = r.lpop(qkey)
        if child_id:
            if isinstance(child_id, bytes):
                child_id = child_id.decode("utf-8")
            _last_parent_pointer = (parents.index(job_id) + 1) % max(1, len(parents))
            return child_id
    return None


def _run_child_thread(child_id: str):
    try:
        import dbs_tasks
        dbs_tasks.process_dbs_child(_child_get(child_id) or {})
    except Exception as e:
        log.exception("[Dispatcher] child %s failed: %s", child_id, e)
    finally:
        _unregister_active(child_id)


def _dispatcher_loop():
    _recover_stale_children()
    while True:
        try:
            while _active_count() < MAX_CONCURRENT_TASKS:
                child_id = _pick_next_child()
                if not child_id:
                    break
                child = _child_get(child_id)
                if not child or child.get("status") != "queued":
                    continue
                child["status"] = "running"
                _child_set(child_id, child)
                _register_active(child_id)
                import threading
                t = threading.Thread(target=_run_child_thread, args=(child_id,), daemon=True, name=f"dbs-child-{child_id[:8]}")
                t.start()
            time.sleep(WORKER_POLL_INTERVAL)
        except Exception as e:
            log.exception("[Dispatcher] loop error: %s", e)
            time.sleep(WORKER_POLL_INTERVAL)


def _now_iso() -> str:
    return datetime.datetime.utcnow().isoformat()


def _normalize_item_fields(it: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(it or {})
    dd, mm, yy = normalize_dob_parts(out.get("dob_day"), out.get("dob_month"), out.get("dob_year"))
    out["dob_day"], out["dob_month"], out["dob_year"] = dd, mm, yy
    idd, imm, iyy = normalize_dob_parts(out.get("issue_day"), out.get("issue_month"), out.get("issue_year"))
    out["issue_day"], out["issue_month"], out["issue_year"] = idd, imm, iyy
    out["certificate_number"] = (out.get("certificate_number") or "").strip()
    out["surname"] = (out.get("surname") or "").strip()
    out["forename"] = (out.get("forename") or "").strip()
    return out


def _create_parent_job(tenant_id: int, user_id: int, items: List[dict], db_job_id: Optional[int], storage_path: Path,
                       org_name: str, emp_fn: str, emp_sn: str, reuse_rows: Optional[List[dict]] = None,
                       old_job_id: Optional[str] = None) -> tuple[str, list[dict]]:
    r = get_redis()
    if not r:
        raise HTTPException(500, "Redis unavailable.")
    job_id = str(uuid.uuid4())
    if reuse_rows is None:
        rows = [{"row": i + 1, "status": "queued",
                 "certificate_number": (it.get("certificate_number", "") if isinstance(it, dict) else ""),
                 "surname": (it.get("surname", "") if isinstance(it, dict) else ""),
                 "forename": (it.get("forename", "") if isinstance(it, dict) else ""),
                 "dob_day": (it.get("dob_day", "") if isinstance(it, dict) else ""),
                 "dob_month": (it.get("dob_month", "") if isinstance(it, dict) else ""),
                 "dob_year": (it.get("dob_year", "") if isinstance(it, dict) else ""),
                 "issue_day": (it.get("issue_day", "") if isinstance(it, dict) else ""),
                 "issue_month": (it.get("issue_month", "") if isinstance(it, dict) else ""),
                 "issue_year": (it.get("issue_year", "") if isinstance(it, dict) else ""),
                 "original_filename": (it.get("original_filename", f"Row {i+1}") if isinstance(it, dict) else f"Row {i+1}"),
                 "pdf_filename": "", "pdf_url": "", "error": "", "notes": ""}
                for i, it in enumerate(items)]
    else:
        rows = reuse_rows
    state = {
        "state": "queued",
        "rows": rows,
        "zip_ready": False,
        "zip_name": "",
        "zip_url": "",
        "checked_date": "",
        "message": "Batch queued — processing starts shortly...",
        "successful": 0,
        "failed": 0,
        "queued_count": sum(1 for r0 in rows if r0.get("status") == "queued"),
        "running_count": 0,
        "parent_total": len(rows),
        "db_job_id": db_job_id,
        "tenant_id": tenant_id,
        "user_id": user_id,
        "storage_path": str(storage_path),
        "organisation_name": org_name,
        "employee_forename": emp_fn,
        "employee_surname": emp_sn,
        "created_at": _now_iso(),
        "source_job_id": old_job_id or "",
    }
    _jset(job_id, state)
    _owner_set(job_id, tenant_id)
    return job_id, rows


def _enqueue_child_tasks(job_id: str, tenant_id: int, user_id: int, storage_path: Path, org_name: str, emp_fn: str, emp_sn: str, items: List[dict]):
    r = get_redis()
    for idx, it in enumerate(items, start=1):
        child_id = str(uuid.uuid4())
        row_number = int(it.get("row_number") or idx)
        payload = {
            "child_id": child_id,
            "job_id": job_id,
            "tenant_id": tenant_id,
            "user_id": user_id,
            "row_number": row_number,
            "storage_path": str(storage_path),
            "organisation_name": org_name,
            "employee_forename": emp_fn,
            "employee_surname": emp_sn,
            "certificate_number": (it.get("certificate_number") or "").strip(),
            "surname": (it.get("surname") or "").strip(),
            "forename": (it.get("forename") or "").strip(),
            "dob_day": (it.get("dob_day") or "").strip(),
            "dob_month": (it.get("dob_month") or "").strip(),
            "dob_year": (it.get("dob_year") or "").strip(),
            "issue_day": (it.get("issue_day") or "").strip(),
            "issue_month": (it.get("issue_month") or "").strip(),
            "issue_year": (it.get("issue_year") or "").strip(),
            "original_filename": it.get("original_filename") or f"Row {idx}",
            "status": "queued",
            "created_at": _now_iso(),
        }
        _child_set(child_id, payload)
        r.rpush(_parent_queue_key(job_id), child_id)


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request, "dashboard_url": APP_DASHBOARD_URL, "login_url": APP_LOGIN_URL})


@app.get("/health")
def health():
    r = get_redis(); ok = False
    try:
        if r: r.ping(); ok = True
    except: pass
    return {"ok": True, "redis": ok, "gemini": bool(GEMINI_CLIENT), "db": bool(os.getenv("DATABASE_URL")), "max_concurrent_tasks": MAX_CONCURRENT_TASKS}

@app.post("/dbs/extract")
async def dbs_extract(request: Request, files: List[UploadFile] = File(...)):
    _auth(request)
    if not files: raise HTTPException(400, "No file uploaded.")
    items: List[Dict] = []; cap = 100
    for file in files[:100]:
        content = await file.read()
        fname = file.filename or ""; filename = fname.lower()
        if len(content) > 25*1024*1024: raise HTTPException(413, f"'{fname}' too large.")
        if filename.endswith(".docx"):
            try:
                from docx import Document
                doc = Document(io.BytesIO(content))
                text = "\n".join([p.text for p in doc.paragraphs if (p.text or "").strip()])
            except Exception as e: raise HTTPException(400, f"DOCX error: {e}")
            fields = extract_fields_from_text(text)
            dob = fields.get("dob") if isinstance(fields.get("dob"), dict) else {}
            issue = fields.get("issue_date") if isinstance(fields.get("issue_date"), dict) else {}
            items.append({"original_filename": fname, "forename": fields.get("forename",""),
                "certificate_number": fields.get("certificate_number",""), "surname": fields.get("surname",""),
                "dob_day": dob.get("dd",""), "dob_month": dob.get("mm",""), "dob_year": dob.get("yyyy",""),
                "issue_day": issue.get("dd",""), "issue_month": issue.get("mm",""), "issue_year": issue.get("yyyy",""),
                "confidence": {"certificate_number": score_cert_number(fields.get("certificate_number","")),
                               "surname": score_surname(fields.get("surname",""),source="docx"),
                               "dob": score_dob(dob.get("dd",""),dob.get("mm",""),dob.get("yyyy",""),source="docx"), "issue_date": 0},
                "source": {"certificate_number":"DOCX","surname":"DOCX","dob":"DOCX","issue_date":""}})
            if len(items) >= cap: break; continue
        if filename.endswith(".webp"):
            try:
                from PIL import Image
                im = Image.open(io.BytesIO(content)).convert("RGB"); buf = io.BytesIO(); im.save(buf, "PNG")
                content = buf.getvalue(); filename = fname.lower().replace(".webp",".png")
            except Exception as e: raise HTTPException(400, f"WEBP error: {e}")
        if filename.endswith(".csv") or filename.endswith(".xlsx"):
            try: rows = parse_csv_rows(content) if filename.endswith(".csv") else parse_xlsx_rows(content)
            except Exception as e: raise HTTPException(400, f"Spreadsheet error: {e}")
            for ri, r in enumerate(rows, 2):
                if len(items) >= cap: break
                items.append({"original_filename": f"{fname} (Row {ri})", "forename": r.get("forename",""),
                    "certificate_number": r.get("certificate_number",""), "surname": r.get("surname",""),
                    "dob_day": r.get("dob_day",""), "dob_month": r.get("dob_month",""), "dob_year": r.get("dob_year",""),
                    "issue_day": r.get("issue_day",""), "issue_month": r.get("issue_month",""), "issue_year": r.get("issue_year",""),
                    "confidence": {"certificate_number": score_cert_number(r.get("certificate_number","")),
                                   "surname": score_surname(r.get("surname",""),source="Spreadsheet"),
                                   "dob": score_dob(r.get("dob_day",""),r.get("dob_month",""),r.get("dob_year",""),source="Spreadsheet"), "issue_date": 0},
                    "source": {"certificate_number":"Spreadsheet","surname":"Spreadsheet","dob":"Spreadsheet","issue_date":"Spreadsheet"}})
            if len(items) >= cap: break; continue
        # PDF / image
        fields = {"certificate_number": None, "surname": None, "dob": None, "issue_date": None}
        source = {"certificate_number": "", "surname": "", "dob": "", "issue_date": ""}
        vision = {}
        if filename.endswith(".pdf"):
            text = extract_text_from_pdf(content)
            if text and len(text) > 60:
                fields = extract_fields_from_text(text)
                for k in ("certificate_number","surname","dob","issue_date"):
                    if fields.get(k): source[k] = "PDF text"
            if not fields.get("certificate_number") or not fields.get("surname") or not fields.get("dob"):
                try:
                    imgs = pdf_to_images_bytes(content, max_pages=1, dpi=240)
                    vision = gemini_vision_extract_images([(b,"image/png") for b in imgs])
                    for k in ("certificate_number","surname","forename","dob","issue_date"):
                        if not fields.get(k) and vision.get(k): fields[k] = vision.get(k); source[k] = "Image scan"
                except Exception as e: raise HTTPException(400, str(e))
        else:
            is_png = content[:8] == b"\x89PNG\r\n\x1a\n"; mime = "image/png" if is_png else "image/jpeg"
            vision = gemini_vision_extract_images([(content, mime)])
            fields = {k: vision.get(k) for k in ("certificate_number","surname","forename","dob","issue_date")}
            for k in ("certificate_number","surname","dob","issue_date"):
                if fields.get(k): source[k] = "Image scan"
        dob = fields.get("dob") if isinstance(fields.get("dob"), dict) else {}
        issue = fields.get("issue_date") if isinstance(fields.get("issue_date"), dict) else {}
        vconf = (vision.get("confidence") if isinstance(vision, dict) else {}) or {}
        items.append({"original_filename": fname, "forename": fields.get("forename",""),
            "certificate_number": fields.get("certificate_number",""), "surname": fields.get("surname",""),
            "dob_day": dob.get("dd",""), "dob_month": dob.get("mm",""), "dob_year": dob.get("yyyy",""),
            "issue_day": issue.get("dd",""), "issue_month": issue.get("mm",""), "issue_year": issue.get("yyyy",""),
            "confidence": {"certificate_number": int(vconf.get("certificate_number") or score_cert_number(fields.get("certificate_number",""))) ,
                           "surname": int(vconf.get("surname") or score_surname(fields.get("surname",""),source=source.get("surname",""))),
                           "dob": int(vconf.get("dob") or score_dob(dob.get("dd",""),dob.get("mm",""),dob.get("yyyy",""),source=source.get("dob",""))),
                           "issue_date": int(vconf.get("issue_date") or 0)},
            "source": source})
        if len(items) >= cap: break
    resp = {"items": items}
    if len(items) >= cap and len(files) > 0: resp["notice"] = "Row limit reached (100)."
    return JSONResponse(resp)

@app.post("/dbs/export/extract")
async def export_extract(request: Request):
    _auth(request)
    data = await request.json(); fmt = (data.get("format") or "xlsx").lower()
    rows = _export_rows_extract(data.get("items") or [])
    cols = ["Forename","Surname","Certificate Number","DOB","Issue Date","PDF Filename","Notes"]
    if fmt == "csv":
        return Response(_csv_bytes(rows,cols), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=extract.csv"})
    return Response(_xlsx_bytes(rows,cols), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":"attachment; filename=extract.xlsx"})

@app.post("/dbs/export/results")
async def export_results(request: Request):
    _auth(request)
    data = await request.json(); fmt = (data.get("format") or "xlsx").lower()
    rows = _export_rows_results(data)
    cols = ["Forename","Surname","Certificate Number","DOB","Issue Date","Status","Checked Date","PDF Filename","Notes"]
    if fmt == "csv":
        return Response(_csv_bytes(rows,cols), media_type="text/csv", headers={"Content-Disposition":"attachment; filename=results.csv"})
    return Response(_xlsx_bytes(rows,cols), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition":"attachment; filename=results.xlsx"})

@app.post("/dbs/run")
async def dbs_run(request: Request):
    ctx = _auth(request)
    tenant_id = ctx["tenant_id"]; user_id = ctx["user_id"]
    payload: Dict = {}
    ctype = (request.headers.get("content-type") or "").lower()
    if "application/json" in ctype:
        try: payload = await request.json()
        except: payload = {}
    else:
        form = await request.form(); payload = dict(form)
    org_name = (payload.get("organisation_name") or payload.get("org_name") or "").strip()
    emp_fn = (payload.get("employee_forename") or payload.get("forename") or "").strip()
    emp_sn = (payload.get("employee_surname") or payload.get("surname_user") or payload.get("surname") or "").strip()
    if not (org_name and emp_fn and emp_sn):
        raise HTTPException(400, "Organisation/Forename/Surname is required.")
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        raise HTTPException(400, "No items provided.")
    items = [_normalize_item_fields(it) for it in items[:100]]
    try:
        import db
        tokens = db.get_tenant_tokens_remaining(tenant_id)
        if tokens == 0: raise HTTPException(402, "No tokens remaining.")
        if 0 < tokens < len(items): items = items[:tokens]
    except HTTPException: raise
    except Exception as e: log.warning("[Run] Token check skipped: %s", e)
    storage_path = _storage(tenant_id, user_id, str(uuid.uuid4()))
    db_job_id = None
    try:
        import db
        db_job_id = db.create_job_record(tenant_id=tenant_id, user_id=user_id, total_items=len(items))
    except Exception as e: log.warning("[Run] DB record failed: %s", e)
    job_id, rows = _create_parent_job(tenant_id, user_id, items, db_job_id, storage_path, org_name, emp_fn, emp_sn)
    real_storage = _storage(tenant_id, user_id, job_id)
    if storage_path != real_storage and storage_path.exists() and not any(storage_path.iterdir()):
        try: storage_path.rmdir()
        except Exception: pass
    st = _jget(job_id) or {}
    st["storage_path"] = str(real_storage)
    _jset(job_id, st)
    _enqueue_child_tasks(job_id, tenant_id, user_id, real_storage, org_name, emp_fn, emp_sn, items)
    return JSONResponse({"job_id": job_id, "mode": "bulk", "status_url": f"/dbs/status/{job_id}", "rows": rows, "queued": True})


@app.get("/dbs/status/{job_id}")
async def dbs_status(job_id: str, request: Request):
    _auth(request)
    state = _jget(job_id)
    if not state: raise HTTPException(404, "Job expired or not found.")
    rows = state.get("rows") or []
    done = sum(1 for r in rows if r.get("status") in ("clear", "needs_review", "portal_unavailable", "failed"))
    running = sum(1 for r in rows if r.get("status") == "running")
    queued = sum(1 for r in rows if r.get("status") == "queued")
    return JSONResponse({"job_id": job_id, "mode": "bulk", "state": state.get("state","queued"), "checked_date": state.get("checked_date",""), "running": {"done": done, "total": len(rows) or 1}, "rows": rows, "zip_ready": bool(state.get("zip_ready")), "zip_name": state.get("zip_name",""), "zip_url": state.get("zip_url",""), "message": state.get("message",""), "successful": state.get("successful",0), "failed": state.get("failed",0), "running_count": running, "queued_count": queued})


@app.get("/dbs/download/{job_id}/{name}")
async def dbs_download(job_id: str, name: str, request: Request):
    ctx = _auth(request); tenant_id = ctx["tenant_id"]
    job_tenant = _owner_get(job_id)
    if job_tenant is not None and job_tenant != tenant_id:
        raise HTTPException(403, "Access denied.")
    tenant_root = STORAGE_ROOT / str(tenant_id); file_path = None
    for p in tenant_root.rglob(name):
        if job_id in str(p): file_path = p; break
    if not file_path:
        state = _jget(job_id) or {}
        source_job_id = state.get("source_job_id") or ""
        for p in tenant_root.rglob(name):
            if source_job_id and source_job_id in str(p):
                file_path = p; break
    if not file_path:
        matches = list(tenant_root.rglob(name))
        if len(matches) == 1:
            file_path = matches[0]
    if not file_path or not file_path.exists():
        raise HTTPException(404, "Download expired or not found.")
    bg = None
    if name.lower().endswith(".zip"):
        sp = STORAGE_ROOT / str(tenant_id) / str(ctx["user_id"]) / job_id
        bg = BackgroundTask(_cleanup, sp, job_id)
    return FileResponse(str(file_path), filename=name, media_type="application/octet-stream", background=bg)


def _cleanup(storage_path: Path, job_id: str):
    import time as t; t.sleep(5)
    try:
        if storage_path.exists(): shutil.rmtree(storage_path, ignore_errors=True)
    except Exception as e: log.warning("[Cleanup] %s", e)
    r = get_redis()
    if r:
        try:
            r.delete(_job_key(job_id)); r.delete(_owner_key(job_id)); r.delete(_parent_queue_key(job_id))
            for key in r.scan_iter(match=f"{CHILD_PREFIX}*"):
                child = _job_to_dict(r.get(key))
                if child and child.get("job_id") == job_id:
                    r.delete(key)
        except Exception: pass


@app.post("/dbs/rerun")
async def dbs_rerun(request: Request):
    ctx = _auth(request)
    tenant_id = ctx["tenant_id"]; user_id = ctx["user_id"]
    try: payload = await request.json()
    except: raise HTTPException(400, "Invalid JSON.")
    old_job_id = (payload.get("job_id") or "").strip()
    dirty_items = payload.get("items") or []
    org_name  = (payload.get("organisation_name") or "").strip()
    emp_fn    = (payload.get("employee_forename") or "").strip()
    emp_sn    = (payload.get("employee_surname") or "").strip()
    if not old_job_id: raise HTTPException(400, "job_id required.")
    if not dirty_items: raise HTTPException(400, "No items for rerun.")
    job_tenant = _owner_get(old_job_id)
    if job_tenant is not None and job_tenant != tenant_id:
        raise HTTPException(403, "Access denied.")
    old_state = _jget(old_job_id)
    if not old_state: raise HTTPException(404, "Original job expired. Please run fresh.")
    old_rows = old_state.get("rows") or []
    storage_path = STORAGE_ROOT / str(tenant_id) / str(user_id) / old_job_id
    old_db_job_id = old_state.get("db_job_id")
    try:
        import db
        tokens = db.get_tenant_tokens_remaining(tenant_id)
        if tokens == 0: raise HTTPException(402, "No tokens remaining.")
        if 0 < tokens < len(dirty_items): dirty_items = dirty_items[:tokens]
    except HTTPException: raise
    except Exception as e: log.warning("[Rerun] Token check skipped: %s", e)
    db_job_id = None
    try:
        import db
        db_job_id = db.create_job_record(tenant_id=tenant_id, user_id=user_id, total_items=len(dirty_items))
    except Exception as e: log.warning("[Rerun] DB record: %s", e)
    dirty_items = [_normalize_item_fields(it) for it in dirty_items]
    dirty_map = {int(it.get("row_number") or 0): it for it in dirty_items}
    merged_rows = []
    waived_count = 0
    updated_old_rows = []
    for r0 in old_rows:
        row_no = int(r0.get("row") or 0)
        current_old = dict(r0)
        if row_no in dirty_map:
            if _row_is_billable(current_old):
                current_old["billing_waived"] = True
                current_old["notes"] = ((current_old.get("notes") or "") + " Waived after user edit/rerun.").strip()
                waived_count += 1
            updated_old_rows.append(current_old)
            merged_rows.append({**current_old, "status": "queued", "certificate_number": (dirty_map[row_no].get("certificate_number") or "").strip(), "surname": (dirty_map[row_no].get("surname") or "").strip(), "forename": (dirty_map[row_no].get("forename") or "").strip(), "dob_day": (dirty_map[row_no].get("dob_day") or "").strip(), "dob_month": (dirty_map[row_no].get("dob_month") or "").strip(), "dob_year": (dirty_map[row_no].get("dob_year") or "").strip(), "issue_day": (dirty_map[row_no].get("issue_day") or "").strip(), "issue_month": (dirty_map[row_no].get("issue_month") or "").strip(), "issue_year": (dirty_map[row_no].get("issue_year") or "").strip(), "pdf_filename": "", "pdf_url": "", "error": "", "notes": "", "billing_waived": False})
        else:
            updated_old_rows.append(current_old)
            merged_rows.append(current_old)
    if waived_count > 0:
        old_state["rows"] = updated_old_rows
        old_state["successful"] = sum(1 for row in updated_old_rows if _row_is_billable(row))
        old_state["failed"] = sum(1 for row in updated_old_rows if (row.get("status") in ("portal_unavailable", "failed")))
        old_state["message"] = f"{waived_count} earlier billed row(s) waived after user edits."
        _jset(old_job_id, old_state)
        try:
            import db
            db.reverse_usage(tenant_id=tenant_id, user_id=user_id, db_job_id=old_db_job_id, reversed_outputs=waived_count)
        except Exception as e:
            log.warning("[Rerun] reverse_usage failed: %s", e)
    job_id, rows = _create_parent_job(tenant_id, user_id, dirty_items, db_job_id, storage_path, org_name, emp_fn, emp_sn, reuse_rows=merged_rows, old_job_id=old_job_id)
    _enqueue_child_tasks(job_id, tenant_id, user_id, storage_path, org_name, emp_fn, emp_sn, dirty_items)
    return JSONResponse({"job_id": job_id, "status_url": f"/dbs/status/{job_id}", "rows": rows, "queued": True})
